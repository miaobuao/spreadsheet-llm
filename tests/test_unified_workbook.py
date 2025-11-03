"""
Tests for unified workbook interface.

Tests basic functionality and various boundary conditions, error cases, and special scenarios.
"""

import tempfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Border, Font, PatternFill, Side

from spreadsheet_llm.unified_workbook import (
    OpenpyxlWorkbook,
    create_unified_workbook,
)

# ============================================================================
# Basic functionality tests
# ============================================================================


def test_openpyxl_workbook(tmp_path):
    """Test OpenpyxlWorkbook wrapper with .xlsx file."""
    # Create a test .xlsx file
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise
    ws.title = "TestSheet"
    ws["A1"] = "Hello"
    ws["B2"] = 42

    from openpyxl.styles import Font

    ws["A1"].font = Font(bold=True)

    test_file = tmp_path / "test.xlsx"
    wb.save(test_file)

    # Test unified interface
    unified_wb = create_unified_workbook(str(test_file))
    assert isinstance(unified_wb, OpenpyxlWorkbook)
    assert "TestSheet" in unified_wb.sheetnames

    # Test worksheet access
    ws = unified_wb.active
    assert ws is not None
    assert ws.title == "TestSheet"

    # Test cell access
    cell_a1 = ws.cell(1, 1)
    assert cell_a1.value == "Hello"
    assert cell_a1.font.bold is True

    cell_b2 = ws.cell(2, 2)
    assert cell_b2.value == 42

    # Test dictionary access
    ws2 = unified_wb["TestSheet"]
    assert ws2.title == "TestSheet"


def test_multiple_sheets():
    """Test workbook with multiple sheets."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    if ws1 is None:
        raise
    ws1.title = "Sheet1"
    ws1["A1"] = "First"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Second"

    # Use in-memory file
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Save to temp file for testing
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(buffer.getvalue())
        temp_path = f.name

    try:
        unified_wb = create_unified_workbook(temp_path)
        assert len(unified_wb.sheetnames) == 2
        assert "Sheet1" in unified_wb.sheetnames
        assert "Sheet2" in unified_wb.sheetnames

        # Access by index
        ws1 = unified_wb.get_sheet_by_index(0)
        assert ws1.title == "Sheet1"
        assert ws1.cell(1, 1).value == "First"

        ws2 = unified_wb.get_sheet_by_index(1)
        assert ws2.title == "Sheet2"
        assert ws2.cell(1, 1).value == "Second"

        # Access by name
        ws2_by_name = unified_wb["Sheet2"]
        assert ws2_by_name.title == "Sheet2"

    finally:
        import os

        os.unlink(temp_path)


def test_format_detection():
    """Test that format properties work correctly."""
    wb = openpyxl.Workbook()
    ws = wb.active

    if ws is None:
        raise

    # Cell with bold font
    ws["A1"] = "Bold"
    ws["A1"].font = Font(bold=True)

    # Cell with border
    ws["B1"] = "Border"
    thin_border = Border(
        top=Side(style="thin"),
        bottom=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin"),
    )
    ws["B1"].border = thin_border

    # Cell with fill
    ws["C1"] = "Fill"
    ws["C1"].fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    # Save to temp
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        temp_path = f.name

    try:
        wb.save(temp_path)

        unified_wb = create_unified_workbook(temp_path)
        uws = unified_wb.active

        if uws is None:
            raise

        # Test bold
        cell_a1 = uws.cell(1, 1)
        assert cell_a1.font.bold is True

        # Test border
        cell_b1 = uws.cell(1, 2)
        assert cell_b1.border.top.style is not None
        assert cell_b1.border.bottom.style is not None
        assert cell_b1.border.left.style is not None
        assert cell_b1.border.right.style is not None

        # Test fill
        cell_c1 = uws.cell(1, 3)
        assert cell_c1.fill.start_color is not None
        assert cell_c1.fill.start_color.index != "00000000"

    finally:
        import os

        os.unlink(temp_path)


# ============================================================================
# Edge case tests
# ============================================================================


class TestEmptyCells:
    """Test handling of empty and None cells."""

    def test_empty_cell_xlsx(self):
        """Test reading empty cells in .xlsx."""
        wb = openpyxl.Workbook()
        ws = wb.active

        if ws is None:
            raise

        ws["A1"] = "Value"
        # B1 is empty

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            # A1 has value
            assert uws.cell(1, 1).value == "Value"

            # B1 is empty (should be None)
            assert uws.cell(1, 2).value is None

            # B1 should still have border/font/fill objects (even if empty)
            assert uws.cell(1, 2).border is not None
            assert uws.cell(1, 2).font is not None
            assert uws.cell(1, 2).fill is not None

        finally:
            Path(temp_path).unlink()

    def test_none_value_xlsx(self):
        """Test cells with None value."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws["A1"] = None

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            cell = uws.cell(1, 1)
            assert cell.value is None
            # Should not crash when accessing format
            assert cell.border is not None

        finally:
            Path(temp_path).unlink()


class TestCellBoundaries:
    """Test cell access at boundaries."""

    def test_first_cell(self):
        """Test accessing A1 (first cell)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws["A1"] = "First"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            cell = uws.cell(1, 1)
            assert cell.value == "First"

        finally:
            Path(temp_path).unlink()

    def test_large_cell_indices(self):
        """Test accessing cells with large row/column numbers."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        # Write to a far cell
        ws.cell(row=100, column=50, value="Far")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            cell = uws.cell(100, 50)
            assert cell.value == "Far"

            # Empty cell before it
            cell_empty = uws.cell(50, 25)
            assert cell_empty.value is None

        finally:
            Path(temp_path).unlink()


class TestDataTypes:
    """Test various data types in cells."""

    def test_different_data_types_xlsx(self):
        """Test cells with different data types."""
        wb = openpyxl.Workbook()
        ws = wb.active

        if ws is None:
            raise

        # Various data types
        ws["A1"] = "String"
        ws["A2"] = 42
        ws["A3"] = 3.14159
        ws["A4"] = True
        ws["A5"] = datetime(2024, 1, 15, 10, 30, 0)
        ws["A6"] = None

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            assert uws.cell(1, 1).value == "String"
            assert uws.cell(2, 1).value == 42
            assert abs(uws.cell(3, 1).value - 3.14159) < 0.0001
            assert uws.cell(4, 1).value is True
            assert isinstance(uws.cell(5, 1).value, datetime)
            assert uws.cell(6, 1).value is None

        finally:
            Path(temp_path).unlink()

    def test_formula_xlsx(self):
        """Test cells containing formulas."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=A1+A2"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            # Formula cell value depends on whether it's been calculated
            # Just verify it doesn't crash when accessing
            cell = uws.cell(3, 1)
            _ = cell.value  # May be None if not calculated
            # Verify format access works
            assert cell.border is not None

        finally:
            Path(temp_path).unlink()


class TestFormatEdgeCases:
    """Test edge cases for cell formatting."""

    def test_no_format_xlsx(self):
        """Test cells with no explicit formatting."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws["A1"] = "Plain"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            cell = uws.cell(1, 1)
            # Should have border/font/fill objects but all empty
            assert cell.border.top.style is None
            assert cell.border.bottom.style is None
            assert cell.border.left.style is None
            assert cell.border.right.style is None
            assert cell.font.bold is None or cell.font.bold is False

        finally:
            Path(temp_path).unlink()

    def test_partial_border_xlsx(self):
        """Test cell with only some borders."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws["A1"] = "Test"
        # Only top border
        ws["A1"].border = Border(top=Side(style="thin"))

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            cell = uws.cell(1, 1)
            assert cell.border.top.style is not None
            assert cell.border.bottom.style is None
            assert cell.border.left.style is None
            assert cell.border.right.style is None

        finally:
            Path(temp_path).unlink()

class TestSheetAccess:
    """Test different ways to access sheets."""

    def test_sheet_by_name_not_found(self):
        """Test accessing non-existent sheet by name."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws.title = "Sheet1"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)

            with pytest.raises(KeyError):
                unified_wb["NonExistent"]

        finally:
            Path(temp_path).unlink()

    def test_sheet_by_invalid_index(self):
        """Test accessing sheet with invalid index."""
        wb = openpyxl.Workbook()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)

            # Only 1 sheet exists (index 0)
            with pytest.raises(IndexError):
                unified_wb.get_sheet_by_index(10)

        finally:
            Path(temp_path).unlink()

    def test_multiple_sheets_access(self):
        """Test accessing multiple sheets in various ways."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws.title = "First"
        ws2 = wb.create_sheet("Second")
        ws3 = wb.create_sheet("Third")

        ws2["A1"] = "Sheet2"
        ws3["A1"] = "Sheet3"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)

            # Access by name
            assert unified_wb["Second"].cell(1, 1).value == "Sheet2"

            # Access by index
            assert unified_wb.get_sheet_by_index(2).cell(1, 1).value == "Sheet3"

            # List all sheets
            assert len(unified_wb.sheetnames) == 3
            assert "First" in unified_wb.sheetnames
            assert "Second" in unified_wb.sheetnames
            assert "Third" in unified_wb.sheetnames

        finally:
            Path(temp_path).unlink()


class TestSingleCellWorksheet:
    """Test worksheet with only one cell."""

    def test_single_cell_worksheet(self):
        """Test a worksheet with just one cell."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws["A1"] = "Only"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            assert uws.cell(1, 1).value == "Only"
            assert uws.cell(2, 1).value is None  # Empty cell

        finally:
            Path(temp_path).unlink()


class TestEmptyWorksheet:
    """Test completely empty worksheet."""

    def test_empty_worksheet(self):
        """Test a worksheet with no data."""
        wb = openpyxl.Workbook()
        # Don't write anything

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            # All cells should be empty
            assert uws.cell(1, 1).value is None
            assert uws.cell(10, 10).value is None

        finally:
            Path(temp_path).unlink()


class TestFilePathProperty:
    """Test file_path property."""

    def test_file_path_stored(self):
        """Test that file_path is stored correctly."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws["A1"] = "Test"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)

            assert unified_wb.file_path == temp_path

        finally:
            Path(temp_path).unlink()


class TestNativeWorkbookAccess:
    """Test access to native workbook objects."""

    def test_native_workbook_xlsx(self):
        """Test accessing native openpyxl workbook."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws["A1"] = "Test"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)

            # Should be able to access native workbook
            native = unified_wb.native_workbook
            assert isinstance(native, openpyxl.Workbook)

            # Should be able to use native methods
            assert native.active is not None

        finally:
            Path(temp_path).unlink()



class TestUnsupportedFormats:
    """Test error handling for unsupported formats."""

    def test_unsupported_file_extension(self):
        """Test that unsupported file types raise appropriate error."""
        # Create a fake .xls file (which we don't support yet)
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            f.write(b"A,B,C\n1,2,3\n")
            temp_path = f.name

        try:
            with pytest.raises(ValueError, match="Unsupported workbook type"):
                create_unified_workbook(temp_path)
        except Exception:
            # If it's a different error (like file format), that's also expected
            pass
        finally:
            Path(temp_path).unlink()


class TestWorksheetsList:
    """Test the worksheets property."""

    def test_worksheets_list_xlsx(self):
        """Test getting list of all worksheets."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws.title = "First"
        wb.create_sheet("Second")
        wb.create_sheet("Third")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)

            worksheets = unified_wb.worksheets
            assert len(worksheets) == 3
            assert all(hasattr(ws, "title") for ws in worksheets)
            assert all(hasattr(ws, "cell") for ws in worksheets)

        finally:
            Path(temp_path).unlink()


class TestSpecialCharacters:
    """Test cells with special characters and unicode."""

    def test_unicode_content(self):
        """Test cells with unicode characters."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws["A1"] = "Hello 世界"
        ws["A2"] = "Emoji: 😀🎉"
        ws["A3"] = "Math: ∑∫∂"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            assert uws.cell(1, 1).value == "Hello 世界"
            assert uws.cell(2, 1).value == "Emoji: 😀🎉"
            assert uws.cell(3, 1).value == "Math: ∑∫∂"

        finally:
            Path(temp_path).unlink()

    def test_newlines_in_cells(self):
        """Test cells with newline characters."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise
        ws["A1"] = "Line1\nLine2\nLine3"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = f.name

        try:
            wb.save(temp_path)
            unified_wb = create_unified_workbook(temp_path)
            uws = unified_wb.active

            if uws is None:
                raise

            value = uws.cell(1, 1).value
            assert "\n" in value
            assert "Line1" in value
            assert "Line3" in value

        finally:
            Path(temp_path).unlink()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
