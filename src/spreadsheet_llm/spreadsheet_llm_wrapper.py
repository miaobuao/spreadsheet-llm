import json
import logging
from typing import Any, Dict, List, NamedTuple, Optional, Tuple

import openpyxl
import pandas as pd
from langchain_core.language_models.chat_models import BaseChatModel
from langchain_core.messages import HumanMessage
from pydantic import BaseModel, Field

from spreadsheet_llm.index_column_converter import IndexColumnConverter
from spreadsheet_llm.sheet_compressor import SheetCompressor

logger = logging.getLogger(__name__)


class AnchorsInfo(NamedTuple):
    """Type definition for anchor information returned by compress_spreadsheet."""

    row_anchors: List[int]
    column_anchors: List[int]
    original_shape: Tuple[int, int]
    compressed_shape: Tuple[int, int]


class CompressionResult(NamedTuple):
    """Type definition for compression result returned by compress_spreadsheet."""

    areas: List[Any]
    compress_dict: Dict[Any, List[str]]
    sheet_compressor: SheetCompressor
    anchors: AnchorsInfo
    compressed_sheet: pd.DataFrame
    sheet_name: str


class CellRangeItem(BaseModel):
    """Simple schema for cell range with optional title."""

    title: Optional[str] = Field(
        None, description="Optional title or label for this range"
    )
    range: str = Field(description="Cell range address (e.g., 'A1:B5', 'C3')")


class CellRangeList(BaseModel):
    """List of cell ranges with reasoning (Chain of Thought)."""

    reasoning: str = Field(
        description="Step-by-step reasoning process for identifying the cell ranges. "
        "Explain what patterns you noticed, how you identified meaningful ranges, "
        "and why you chose these specific ranges."
    )
    items: List[CellRangeItem] = Field(
        description="List of cell ranges found in the spreadsheet"
    )


RECOGNIZE_PROMPT = """Instruction:
Given an inverted index mapping cell content to their locations in a spreadsheet.

Format:
   Each line consists of the cell content and the cell addresses where it appears, separated by '|'.
   Format: content|cell1,cell2,cell3,...
   Example:
   Eagles|B12,B39,B44,B48,B52,B54
   Purple City|D12,J12,F16,D18,E41,E49
   ${Integer}|G14:G15,H15,G16:I17

   Data formats are prefixed with '${}', like '${Integer}' or '${yyyy/mm/dd}'.
   Cells are separated by commas.

Your task is to:
1. First, provide step-by-step reasoning about what patterns you observe in the data
2. Explain how you identify meaningful cell ranges
3. Then, return the identified cell ranges with optional descriptive titles

Use Chain of Thought reasoning to ensure accurate identification.
"""


class SpreadsheetLLMWrapper:

    def read_spreadsheet(self, file):

        wb = openpyxl.load_workbook(file)
        return wb

    # Takes a file, compresses it
    def compress_spreadsheet(
        self, wb, format_aware: bool = False, sheet_name: int | str | None = None
    ):
        """
        Compress spreadsheet using SpreadsheetLLM algorithm.

        Args:
            wb: Openpyxl workbook instance
            format_aware: If True, enables format-aware aggregation in dict output.
                         Groups cells by both value AND data type (e.g., "100 (Integer)").
                         If False (default), groups cells only by value.
            sheet_name: Sheet to process. Can be:
                       - None: use active sheet (default)
                       - int: sheet index (0-based)
                       - str: sheet name (e.g., "Sheet1", "Financial Statements")

        Returns:
            CompressionResult containing areas, compress_dict, sheet_compressor, and anchors
        """
        sheet_compressor = SheetCompressor(format_aware=format_aware)

        # Determine which sheet to process
        if sheet_name is None:
            # Use active sheet
            actual_sheet_name = wb.active.title
            logger.info(
                f"Processing active sheet: '{actual_sheet_name}' (total sheets: {len(wb.sheetnames)})"
            )
        elif isinstance(sheet_name, int):
            actual_sheet_name = wb.sheetnames[sheet_name]
            logger.info(
                f"Processing sheet index {sheet_name}: '{actual_sheet_name}' (total sheets: {len(wb.sheetnames)})"
            )
        else:
            actual_sheet_name = sheet_name
            logger.info(
                f"Processing sheet by name: '{sheet_name}' (total sheets: {len(wb.sheetnames)})"
            )

        # header=None: treat all rows as data, don't auto-generate column names
        sheet = pd.read_excel(
            wb, engine="openpyxl", header=None, sheet_name=actual_sheet_name
        )

        # Reset index and column names to integers
        sheet = sheet.reset_index(drop=True)
        sheet.columns = list(range(len(sheet.columns)))

        original_shape = sheet.shape
        logger.info(f"Original sheet shape: {original_shape} (rows x cols)")
        logger.debug(f"First 5 rows:\n{sheet.head()}")

        # Structural-anchor-based Extraction
        sheet = sheet_compressor.anchor(sheet)
        compressed_shape = sheet.shape
        logger.info(f"After anchor, sheet shape: {compressed_shape} (rows x cols)")

        # Collect anchor information
        row_anchors = list(sheet_compressor.row_candidates)
        column_anchors = list(sheet_compressor.column_candidates)

        anchors = AnchorsInfo(
            row_anchors=row_anchors,
            column_anchors=column_anchors,
            original_shape=original_shape,
            compressed_shape=compressed_shape,
        )
        logger.info(
            f"Anchors found: {len(row_anchors)} rows, {len(column_anchors)} columns"
        )

        # Debug: Print all anchor coordinates
        converter = IndexColumnConverter()
        logger.info("=" * 60)
        logger.info("ALL ANCHOR COORDINATES (for debugging):")
        logger.info("=" * 60)
        logger.info(f"Row anchors (0-indexed): {sorted(row_anchors)}")
        logger.info(f"Row anchors (1-indexed): {[r+1 for r in sorted(row_anchors)]}")
        logger.info(f"Column anchors (0-indexed): {sorted(column_anchors)}")
        col_letters = [converter.parse_colindex(c + 1) for c in sorted(column_anchors)]
        logger.info(f"Column anchors (letters): {col_letters}")
        logger.info("=" * 60)

        # Encoding
        markdown = sheet_compressor.encode(
            wb, sheet
        )  # Paper encodes first then anchors; I chose to do this in reverse
        logger.info(f"Encoded markdown shape: {markdown.shape}")
        logger.debug(f"Markdown columns: {markdown.columns.tolist()}")
        logger.debug(f"First 10 markdown entries:\n{markdown.head(10)}")

        # Data-Format Aggregation
        markdown["Category"] = markdown["Value"].apply(
            lambda x: sheet_compressor.get_category(x)
        )
        category_dict = sheet_compressor.inverted_category(markdown)
        logger.info(f"Categories found: {set(category_dict.values())}")
        logger.info(f"Number of unique values: {len(category_dict)}")

        try:
            areas = sheet_compressor.identical_cell_aggregation(sheet, category_dict)
            logger.info(f"Number of areas identified: {len(areas)}")
            logger.debug("First 10 areas:")
            for i, area in enumerate(areas[:10]):
                logger.debug(f"  Area {i}: {area}")
        except RecursionError:
            logger.error("RecursionError in identical_cell_aggregation")
            return

        # Inverted-index Translation
        compress_dict = sheet_compressor.inverted_index(markdown)
        logger.info(f"Compress dict entries: {len(compress_dict)}")

        return CompressionResult(
            areas=areas,
            compress_dict=compress_dict,
            sheet_compressor=sheet_compressor,
            anchors=anchors,
            compressed_sheet=sheet,
            sheet_name=actual_sheet_name,
        )

    def serialize_areas(self, areas, sheet_compressor):
        """Serialize areas to string representation.

        Args:
            areas: List of area tuples
            sheet_compressor: SheetCompressor instance with row/column mappings

        Returns:
            Serialized string representation of areas
        """
        string = ""
        converter = IndexColumnConverter()
        for i in areas:
            # Map compressed indices back to original indices
            original_row_start = sheet_compressor.row_mapping[i[0][0]]
            original_col_start = sheet_compressor.column_mapping[i[0][1]]
            original_row_end = sheet_compressor.row_mapping[i[1][0]]
            original_col_end = sheet_compressor.column_mapping[i[1][1]]

            string += (
                "("
                + i[2]
                + "|"
                + converter.parse_colindex(original_col_start + 1)
                + str(original_row_start + 1)
                + ":"
                + converter.parse_colindex(original_col_end + 1)
                + str(original_row_end + 1)
                + "), "
            )
        return string

    def write_areas(self, file, areas, sheet_compressor):
        """Write serialized areas to file."""
        string = self.serialize_areas(areas, sheet_compressor)
        with open(file, "w+", encoding="utf-8") as f:
            f.writelines(string)

    def serialize_dict(self, dict):
        """Serialize dictionary to string representation.

        Args:
            dict: Dictionary to serialize

        Returns:
            Serialized string representation of dictionary
        """
        string = ""
        for key, value in dict.items():
            # Skip empty keys
            if not key or str(key).strip() == "":
                continue
            # value is now list[str], join with comma
            value_str = ",".join(value) if isinstance(value, list) else str(value)
            string += str(key) + "|" + value_str + "\n"
        return string

    def write_dict(self, file, dict):
        """Write serialized dictionary to file."""
        string = self.serialize_dict(dict)
        with open(file, "w+", encoding="utf-8") as f:
            f.writelines(string)

    def serialize_mapping(self, sheet_compressor):
        """Serialize coordinate mapping to JSON string.

        Args:
            sheet_compressor: SheetCompressor instance with row/column mappings

        Returns:
            JSON string representation of coordinate mapping (sorted by key)
        """
        mapping = sheet_compressor.get_coordinate_mapping()
        return json.dumps(mapping, indent=2, ensure_ascii=False, sort_keys=True)

    def write_mapping(self, file, sheet_compressor):
        """Write coordinate mapping from compressed to original coordinates as JSON"""
        mapping_str = self.serialize_mapping(sheet_compressor)
        with open(file, "w+", encoding="utf-8") as f:
            f.write(mapping_str)

    def convert_compressed_to_original(
        self, compressed_coord: str, sheet_compressor
    ) -> str:
        """
        Convert compressed coordinate(s) to original coordinate(s).

        Uses the coordinate mapping table from get_coordinate_mapping() for direct lookup.

        Args:
            compressed_coord: Compressed coordinate string, can be:
                - Single cell: "A1", "B5"
                - Range: "A1:B5", "C3:D10"
                - Multiple ranges: "A1,B2:B5,C3"
            sheet_compressor: SheetCompressor instance with coordinate mapping

        Returns:
            Original coordinate string in the same format as input

        Examples:
            "A1" -> "A1" (if A1 maps to A1)
            "B5" -> "C10" (if compressed B5 maps to original C10)
            "A1:B3" -> "A1:C5" (range conversion)
            "A1,B2:B5" -> "A1,C3:C8" (multiple ranges)
        """
        # Get the complete coordinate mapping table
        mapping = sheet_compressor.get_coordinate_mapping()

        def convert_single_cell(cell: str) -> str:
            """Convert a single cell coordinate using the mapping table"""
            cell = cell.strip()

            # Direct lookup in mapping table
            if cell in mapping:
                return mapping[cell]
            else:
                logger.warning(f"Cell {cell} not found in mapping table")
                return cell

        def convert_range(range_str: str) -> str:
            """Convert a cell range (e.g., 'A1:B5')"""
            if ":" in range_str:
                start, end = range_str.split(":")
                return f"{convert_single_cell(start)}:{convert_single_cell(end)}"
            else:
                return convert_single_cell(range_str)

        # Handle multiple ranges separated by commas
        parts = compressed_coord.split(",")
        converted_parts = [convert_range(part.strip()) for part in parts]

        return ",".join(converted_parts)

    def recognize(
        self, compress_dict, model: BaseChatModel, user_prompt: str | None = None
    ) -> CellRangeList:
        """
        Use LLM to extract structured cell ranges from spreadsheet content with Chain of Thought reasoning.

        Returns compressed coordinates. Use recognize_original() to get original coordinates.

        Returns a CellRangeList Pydantic model with:
        - reasoning: Step-by-step Chain of Thought explanation
        - items: List of CellRangeItem objects (each with 'title' and 'range')

        Args:
            compress_dict: Inverted index dictionary from compression
            model: LangChain ChatModel instance (e.g., ChatOpenAI, ChatAnthropic, etc.)
            user_prompt: Optional user instruction. If None, extracts all meaningful ranges.

        Returns:
            CellRangeList Pydantic model with compressed coordinates

        Example:
            >>> from langchain_openai import ChatOpenAI
            >>> model = ChatOpenAI(model="gpt-4o-mini")
            >>> wrapper = SpreadsheetLLMWrapper()
            >>> wb = wrapper.read_spreadsheet("data.xlsx")
            >>> result = wrapper.compress_spreadsheet(wb, format_aware=True)
            >>> recognition = wrapper.recognize(result.compress_dict, model=model)
            >>> print(recognition.reasoning)  # See the LLM's thought process
            >>> for item in recognition.items:
            ...     print(f"{item.title or 'N/A'}: {item.range}")
        """

        # Serialize the compressed data
        dict_str = self.serialize_dict(compress_dict)

        # Build the user message with instruction and input
        user_message = RECOGNIZE_PROMPT

        if user_prompt:
            user_message += f"\nAdditional instructions: {user_prompt}\n"

        user_message += "\nINPUT:\n"
        user_message += dict_str

        logger.info(
            f"Sending structured recognition request to LLM: {model.__class__.__name__}"
        )
        logger.info("Output schema: CellRangeList")
        logger.debug(f"User message length: {len(user_message)} chars")

        try:
            # Create a structured output version of the LLM
            structured_llm = model.with_structured_output(CellRangeList)

            # Call LLM using LangChain (only user message, no system message)
            messages = [
                HumanMessage(content=user_message),
            ]

            response = structured_llm.invoke(messages)
            if isinstance(response, dict):
                response = CellRangeList(**response)
            elif not isinstance(response, CellRangeList):
                response = CellRangeList(**response.model_dump())

            logger.info(
                f"Structured recognition completed. Found {len(response.items)} ranges."
            )
            logger.debug(f"Reasoning: {response.reasoning}")
            return response

        except Exception as e:
            logger.error(f"LLM structured output error: {str(e)}")
            raise

    def recognize_original(
        self,
        compress_dict,
        sheet_compressor,
        model: BaseChatModel,
        user_prompt: str | None = None,
    ) -> CellRangeList:
        """
        Use LLM to extract structured cell ranges and convert to original spreadsheet coordinates.

        This method wraps recognize() and automatically converts compressed coordinates
        to original spreadsheet coordinates using the sheet_compressor mappings.

        Args:
            compress_dict: Inverted index dictionary from compression
            sheet_compressor: SheetCompressor instance with row/column mappings
            model: LangChain ChatModel instance (e.g., ChatOpenAI, ChatAnthropic, etc.)
            user_prompt: Optional user instruction. If None, extracts all meaningful ranges.

        Returns:
            CellRangeList Pydantic model with original coordinates

        Example:
            >>> from langchain_openai import ChatOpenAI
            >>> model = ChatOpenAI(model="gpt-4o-mini")
            >>> wrapper = SpreadsheetLLMWrapper()
            >>> wb = wrapper.read_spreadsheet("data.xlsx")
            >>> result = wrapper.compress_spreadsheet(wb, format_aware=True)
            >>> recognition = wrapper.recognize_original(
            ...     result.compress_dict, result.sheet_compressor, model=model
            ... )
            >>> print(recognition.reasoning)
            >>> for item in recognition.items:
            ...     print(f"{item.title or 'N/A'}: {item.range}")  # Original coordinates
        """
        # Call recognize to get compressed coordinates
        compressed_result = self.recognize(compress_dict, model, user_prompt)

        # Convert all ranges to original coordinates
        logger.info("Converting compressed coordinates to original coordinates")
        converted_items = []
        for item in compressed_result.items:
            original_range = self.convert_compressed_to_original(
                item.range, sheet_compressor
            )
            converted_items.append(
                CellRangeItem(title=item.title, range=original_range)
            )

        logger.debug(f"Converted {len(converted_items)} ranges to original coordinates")

        # Return new CellRangeList with original coordinates
        return CellRangeList(
            reasoning=compressed_result.reasoning, items=converted_items
        )
