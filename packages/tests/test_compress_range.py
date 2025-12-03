"""Test compress_range function."""

import datetime

import pytest
from openpyxl import Workbook

from spreadsheet_llm.range_compressor import compress_range


def create_test_workbook():
    """Create a test workbook with sample data."""
    wb = Workbook()
    ws = wb.active

    if ws is None:
        raise RuntimeError("Failed to create active worksheet.")

    # Add sample data
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["C1"] = "Score"

    ws["A2"] = "Alice"
    ws["B2"] = 25
    ws["C2"] = 95

    ws["A3"] = "Bob"
    ws["B3"] = 30
    ws["C3"] = 87

    ws["A4"] = "Charlie"
    ws["B4"] = 25
    ws["C4"] = 92

    ws["A5"] = "David"
    ws["B5"] = 28
    ws["C5"] = 88

    ws["A10"] = "Summary"
    ws["B10"] = "Total"
    ws["C10"] = 362

    return wb


def create_multitype_workbook():
    """Create a workbook with various data types."""
    wb = Workbook()
    ws = wb.active

    if ws is None:
        raise RuntimeError("Failed to create active worksheet.")

    # String
    ws["A1"] = "Product"
    ws["A2"] = "Apple"
    ws["A3"] = "Banana"

    # Integer
    ws["B1"] = "Quantity"
    ws["B2"] = 100
    ws["B3"] = 200

    # Float
    ws["C1"] = "Price"
    ws["C2"] = 1.99
    ws["C3"] = 2.50

    # Date
    ws["D1"] = "Date"
    ws["D2"] = datetime.datetime(2024, 1, 15)
    ws["D3"] = datetime.datetime(2024, 1, 16)

    # Percentage (as float)
    ws["E1"] = "Discount"
    ws["E2"] = 0.15
    ws["E3"] = 0.20

    return wb


def create_sparse_workbook():
    """Create a workbook with sparse data (many empty cells)."""
    wb = Workbook()
    ws = wb.active

    if ws is None:
        raise RuntimeError("Failed to create active worksheet.")

    # Sparse data with gaps
    ws["A1"] = "Header"
    ws["A5"] = "Data1"
    ws["A10"] = "Data2"

    ws["C3"] = "Isolated"
    ws["E7"] = "Another"

    return wb


def create_multisheet_workbook():
    """Create a workbook with multiple sheets."""
    wb = Workbook()

    # First sheet (active)
    ws1 = wb.active
    if ws1 is None:
        raise RuntimeError("Failed to create active worksheet.")
    ws1.title = "Sales"
    ws1["A1"] = "Q1"
    ws1["B1"] = 1000

    # Second sheet
    ws2 = wb.create_sheet("Expenses")
    ws2["A1"] = "Rent"
    ws2["B1"] = 500

    # Third sheet
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "Total"
    ws3["B1"] = 500

    return wb


def test_compress_range_basic():
    """Test basic compress_range functionality."""
    wb = create_test_workbook()
    ws = wb.active

    # Test with a small range
    result = compress_range(ws, ("A1", "C10"), format_aware=True)

    print("=" * 60)
    print("Test: compress_range with format_aware=True")
    print("=" * 60)
    print(f"Result type: {type(result)}")
    print(f"Number of entries: {len(result)}")
    print("\nFirst 10 entries:")
    for i, (key, cells) in enumerate(result.items()):
        if i >= 10:
            break
        print(f"  {key}: {cells}")

    # Verify result structure
    assert isinstance(result, dict)
    assert len(result) > 0

    # Verify that all cell addresses are absolute coordinates
    for key, cell_list in result.items():
        for cell_str in cell_list:
            # Should contain letters (column) and numbers (row)
            assert any(c.isalpha() for c in cell_str)
            assert any(c.isdigit() for c in cell_str)

    print("\n✓ All tests passed!")


def test_compress_range_single_cell():
    """Test compress_range with single cell."""
    wb = create_test_workbook()
    ws = wb.active

    # Test with a single cell
    result = compress_range(ws, ("A1", "A1"), format_aware=False)

    print("\n" + "=" * 60)
    print("Test: compress_range with single cell")
    print("=" * 60)
    print(f"Number of entries: {len(result)}")
    print(f"Result: {result}")

    assert isinstance(result, dict)
    print("\n✓ Single cell test passed!")


def test_compress_range_format_aware():
    """Test format_aware mode properly categorizes data types."""
    wb = create_multitype_workbook()
    ws = wb.active

    # Test with format_aware=True
    result = compress_range(ws, ("A1", "E3"), format_aware=True)

    # Should have type markers like ${Integer}, ${Float}, ${yyyy/mm/dd}
    assert any("${" in str(key) for key in result.keys()), "Should have type markers"

    # Test with format_aware=False
    result_no_format = compress_range(ws, ("A1", "E3"), format_aware=False)

    # Should not have type markers, just values
    assert not any("${" in str(key) for key in result_no_format.keys())

    print("✓ format_aware mode test passed!")


def test_compress_range_sparse_data():
    """Test compress_range with sparse data (many empty cells)."""
    wb = create_sparse_workbook()
    ws = wb.active

    result = compress_range(ws, ("A1", "E10"), format_aware=False)

    # Should only include non-empty cells
    assert len(result) > 0
    assert "Header" in result
    assert "Data1" in result
    assert "Isolated" in result

    # Empty cells should not appear in result
    total_cells = sum(len(cells) for cells in result.values())
    assert total_cells == 5  # Only 5 non-empty cells

    print("✓ Sparse data test passed!")


def test_compress_range_single_row():
    """Test compress_range with single row."""
    wb = create_test_workbook()
    ws = wb.active

    result = compress_range(ws, ("A1", "C1"), format_aware=False)

    # Should contain header row data
    assert "Name" in result
    assert "Age" in result
    assert "Score" in result

    # Verify coordinates are in row 1
    for cells in result.values():
        for cell in cells:
            assert "1" in cell

    print("✓ Single row test passed!")


def test_compress_range_single_column():
    """Test compress_range with single column."""
    wb = create_test_workbook()
    ws = wb.active

    result = compress_range(ws, ("A1", "A5"), format_aware=False)

    # Should contain column A data
    assert "Name" in result
    assert "Alice" in result
    assert "Bob" in result

    # Verify all coordinates start with A
    for cells in result.values():
        for cell in cells:
            assert cell.startswith("A") or "A" in cell.split(":")[0]

    print("✓ Single column test passed!")


def test_compress_range_offset_range():
    """Test compress_range with range not starting at A1."""
    wb = create_test_workbook()
    ws = wb.active

    # Extract range starting from B2
    result = compress_range(ws, ("B2", "C5"), format_aware=True)

    # Should return absolute coordinates (B2:C5, not A1:B4)
    for cells in result.values():
        for cell_or_range in cells:
            # Extract cell addresses
            if ":" in cell_or_range:
                start, end = cell_or_range.split(":")
                assert start[0] in ["B", "C"]
                assert int(start[1:]) >= 2
            else:
                assert cell_or_range[0] in ["B", "C"]
                assert int(cell_or_range[1:]) >= 2

    print("✓ Offset range test passed!")


def test_compress_range_multisheet_by_name():
    """Test compress_range with specific sheet name."""
    wb = create_multisheet_workbook()

    # Test accessing second sheet by name
    ws = wb["Expenses"]
    result = compress_range(ws, ("A1", "B1"))

    assert "Rent" in result
    assert result["Rent"] == ["A1"]

    print("✓ Multi-sheet by name test passed!")


def test_compress_range_multisheet_by_index():
    """Test compress_range with sheet index."""
    wb = create_multisheet_workbook()

    # Test accessing first sheet (index 0)
    ws1 = wb.worksheets[0]
    result = compress_range(ws1, ("A1", "B1"))

    assert "Q1" in result

    # Test accessing second sheet (index 1)
    ws2 = wb.worksheets[1]
    result2 = compress_range(ws2, ("A1", "B1"))

    assert "Rent" in result2

    print("✓ Multi-sheet by index test passed!")


def test_compress_range_invalid_range():
    """Test compress_range with invalid range format."""
    wb = create_test_workbook()
    ws = wb.active

    # Invalid cell format
    with pytest.raises(ValueError):
        compress_range(ws, ("Invalid", "A1"))

    with pytest.raises(ValueError):
        compress_range(ws, ("A1", "Invalid"))

    print("✓ Invalid range test passed!")


def test_compress_range_absolute_coordinates():
    """Test that returned coordinates are absolute (relative to sheet)."""
    wb = create_test_workbook()

    if wb.active is None:
        raise RuntimeError("Failed to create active worksheet.")

    ws = wb.active

    # Extract a range starting from D5
    ws["D5"] = "Test1"
    ws["E6"] = "Test2"

    result = compress_range(ws, ("D5", "E6"), format_aware=False)

    # Coordinates should be D5, E6 (absolute), not A1, B2 (relative to range)
    assert "Test1" in result
    assert result["Test1"] == ["D5"]
    assert "Test2" in result
    assert result["Test2"] == ["E6"]

    print("✓ Absolute coordinates test passed!")


def test_compress_range_empty_cells_skipped():
    """Test that empty cells are not included in result."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create active worksheet.")

    # Create pattern: Value, Empty, Value, Empty
    ws["A1"] = "Header"
    # A2 is empty
    ws["A3"] = "Data"
    # A4 is empty

    result = compress_range(ws, ("A1", "A4"), format_aware=False)

    # Should only have 2 entries
    assert len(result) == 2
    assert "Header" in result
    assert "Data" in result

    print("✓ Empty cells skipped test passed!")


def test_compress_range_duplicate_values():
    """Test compression of duplicate values."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create active worksheet.")

    # Multiple cells with same value
    ws["A1"] = "Duplicate"
    ws["A2"] = "Duplicate"
    ws["A3"] = "Duplicate"
    ws["B1"] = "Duplicate"

    result = compress_range(ws, ("A1", "B3"), format_aware=False)

    # Should combine duplicate cells
    assert "Duplicate" in result
    assert len(result["Duplicate"]) > 0

    # Should contain range notation (cells are combined)
    duplicate_cells = ",".join(result["Duplicate"])
    # Either uses range notation or lists all cells
    assert "A1" in duplicate_cells or ":" in duplicate_cells

    print("✓ Duplicate values test passed!")


def test_compress_range_large_range():
    """Test compress_range with larger range."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create active worksheet.")

    # Fill 20x10 range with data
    for row in range(1, 21):
        for col in range(1, 11):
            ws.cell(row=row, column=col, value=f"R{row}C{col}")

    result = compress_range(ws, ("A1", "J20"), format_aware=False)

    # Should have 200 unique values
    assert len(result) == 200

    # Verify some samples
    assert "R1C1" in result
    assert "R20C10" in result

    print("✓ Large range test passed!")


if __name__ == "__main__":
    test_compress_range_basic()
    test_compress_range_single_cell()
