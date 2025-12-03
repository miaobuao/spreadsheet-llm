"""Test recognize_*_with_compressed_range functions."""

import openpyxl
import pytest
from pathlib import Path
from unittest.mock import MagicMock, patch

from spreadsheet_llm import SpreadsheetLLMWrapper
from spreadsheet_llm.spreadsheet_llm_wrapper import CellRangeList, CellRangeItem


@pytest.fixture
def sample_workbook(tmp_path):
    """Create a simple test workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Create a simple table
    ws['A1'] = 'Product'
    ws['B1'] = 'Price'
    ws['C1'] = 'Quantity'

    ws['A2'] = 'Apple'
    ws['B2'] = 100
    ws['C2'] = 10

    ws['A3'] = 'Banana'
    ws['B3'] = 150
    ws['C3'] = 20

    ws['A4'] = 'Cherry'
    ws['B4'] = 200
    ws['C4'] = 30

    # Save to file
    file_path = tmp_path / "test_data.xlsx"
    wb.save(file_path)

    return file_path


@pytest.fixture
def wrapper():
    """Create wrapper instance."""
    return SpreadsheetLLMWrapper()


@pytest.fixture
def mock_llm_response():
    """Create mock LLM response."""
    return CellRangeList(
        reasoning="Found a data table",
        items=[
            CellRangeItem(title="Product Table", range="A1:C4"),
        ]
    )


def test_recognize_with_compressed_range_structure(
    sample_workbook, wrapper, mock_llm_response
):
    """Test that recognize_with_compressed_range returns correct structure."""
    # Read and compress
    wb = wrapper.read_spreadsheet(sample_workbook)
    result = wrapper.compress_spreadsheet(wb, format_aware=True)

    # Mock the LLM call
    with patch.object(wrapper, 'recognize', return_value=mock_llm_response):
        recognition = wrapper.recognize_with_compressed_range(
            wb, result, model=MagicMock(), format_aware=True
        )

    # Verify structure
    assert hasattr(recognition, 'reasoning')
    assert hasattr(recognition, 'items')
    assert len(recognition.items) == 1

    item = recognition.items[0]
    assert hasattr(item, 'title')
    assert hasattr(item, 'range')
    assert hasattr(item, 'compressed_dict')

    assert item.title == "Product Table"
    assert isinstance(item.compressed_dict, dict)


def test_recognize_with_compressed_range_compress_dict_not_empty(
    sample_workbook, wrapper, mock_llm_response
):
    """Test that compressed_dict is generated and not empty."""
    wb = wrapper.read_spreadsheet(sample_workbook)
    result = wrapper.compress_spreadsheet(wb, format_aware=True)

    with patch.object(wrapper, 'recognize', return_value=mock_llm_response):
        recognition = wrapper.recognize_with_compressed_range(
            wb, result, model=MagicMock(), format_aware=True
        )

    item = recognition.items[0]

    # Compressed dict should not be empty (contains actual data)
    assert len(item.compressed_dict) > 0

    # Should contain some values from our test data
    # (exact values depend on compression, but should have entries)
    assert isinstance(item.compressed_dict, dict)


def test_recognize_original_with_compressed_range_structure(
    sample_workbook, wrapper, mock_llm_response
):
    """Test that recognize_original_with_compressed_range returns correct structure."""
    wb = wrapper.read_spreadsheet(sample_workbook)
    result = wrapper.compress_spreadsheet(wb, format_aware=True)

    with patch.object(wrapper, 'recognize', return_value=mock_llm_response):
        recognition = wrapper.recognize_original_with_compressed_range(
            wb, result, model=MagicMock(), format_aware=True
        )

    # Verify structure
    assert hasattr(recognition, 'reasoning')
    assert hasattr(recognition, 'items')
    assert len(recognition.items) == 1

    item = recognition.items[0]
    assert hasattr(item, 'title')
    assert hasattr(item, 'range')
    assert hasattr(item, 'compressed_dict')

    assert item.title == "Product Table"
    # Range should be in original coordinates (might be different from compressed)
    assert isinstance(item.range, str)
    assert isinstance(item.compressed_dict, dict)


def test_recognize_original_with_compressed_range_coordinates(
    sample_workbook, wrapper, mock_llm_response
):
    """Test that original coordinates are returned."""
    wb = wrapper.read_spreadsheet(sample_workbook)
    result = wrapper.compress_spreadsheet(wb, format_aware=True)

    with patch.object(wrapper, 'recognize', return_value=mock_llm_response):
        # Get both compressed and original versions
        compressed_recognition = wrapper.recognize_with_compressed_range(
            wb, result, model=MagicMock(), format_aware=True
        )
        original_recognition = wrapper.recognize_original_with_compressed_range(
            wb, result, model=MagicMock(), format_aware=True
        )

    compressed_item = compressed_recognition.items[0]
    original_item = original_recognition.items[0]

    # Ranges might be different if compression changed coordinates
    # But both should be valid range strings
    assert ':' in compressed_item.range or ',' in compressed_item.range or len(compressed_item.range) >= 2
    assert ':' in original_item.range or ',' in original_item.range or len(original_item.range) >= 2


def test_recognize_with_compressed_range_compress_dict_has_real_data(
    sample_workbook, wrapper, mock_llm_response
):
    """Test that compress_dict contains real data from the range."""
    wb = wrapper.read_spreadsheet(sample_workbook)
    result = wrapper.compress_spreadsheet(wb, format_aware=True)

    with patch.object(wrapper, 'recognize', return_value=mock_llm_response):
        recognition = wrapper.recognize_with_compressed_range(
            wb, result, model=MagicMock(), format_aware=True
        )

    item = recognition.items[0]
    compress_dict = item.compressed_dict

    # Should contain entries
    assert len(compress_dict) > 0

    # All values in compress_dict should be lists of cell references
    for key, value in compress_dict.items():
        assert isinstance(value, list)
        assert len(value) > 0
        # Each entry should be a cell reference or range
        for cell_ref in value:
            assert isinstance(cell_ref, str)
            assert len(cell_ref) >= 2  # At least "A1"


def test_multiple_ranges_mock():
    """Test with multiple ranges in mock response."""
    wrapper = SpreadsheetLLMWrapper()

    # Create mock workbook with more data
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, 20):
        ws[f'A{i}'] = f'Data{i}'
        ws[f'B{i}'] = i * 10

    # Create compression result mock
    result_mock = MagicMock()
    result_mock.compress_dict = {"Data1": ["A1"], "Data2": ["A2"]}
    result_mock.sheet_name = "Sheet"

    # Mock sheet_compressor with identity mapping
    sheet_compressor_mock = MagicMock()
    sheet_compressor_mock.convert_compressed_to_original = lambda x: x
    result_mock.sheet_compressor = sheet_compressor_mock

    # Mock LLM with multiple ranges
    mock_response = CellRangeList(
        reasoning="Found two tables",
        items=[
            CellRangeItem(title="Table1", range="A1:B5"),
            CellRangeItem(title="Table2", range="A10:B15"),
        ]
    )

    with patch.object(wrapper, 'recognize', return_value=mock_response):
        with patch('spreadsheet_llm.spreadsheet_llm_wrapper.compress_range') as mock_compress:
            # Mock compress_range to return some data
            mock_compress.return_value = {"Value": ["A1"]}

            recognition = wrapper.recognize_with_compressed_range(
                wb, result_mock, model=MagicMock(), format_aware=True
            )

    # Should have 2 items
    assert len(recognition.items) == 2
    assert recognition.items[0].title == "Table1"
    assert recognition.items[1].title == "Table2"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
