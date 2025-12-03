"""
DSBench Table Recognition Script

This script processes Excel files from the DSBench dataset and identifies
table regions using SpreadsheetLLM's compression and LLM-based recognition.

Features:
- Batch processing of Excel files from a directory
- Rich logging with progress tracking
- LLM-based table region identification
- JSON output with detailed recognition results
- Caching support to avoid reprocessing

Usage:
    python tests/test_dsbench_recognition.py
"""

import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from dotenv import load_dotenv

# For coloring cells
from openpyxl.styles import PatternFill
from rich.console import Console
from rich.logging import RichHandler
from rich.progress import (
    BarColumn,
    Progress,
    SpinnerColumn,
    TaskProgressColumn,
    TextColumn,
)
from rich.table import Table

from spreadsheet_llm import SpreadsheetLLMWrapper
from spreadsheet_llm.cell_range_utils import col_to_index, parse_excel_range

load_dotenv()

# Initialize Rich console
console = Console()

# Configuration
DSBENCH_DIR = Path(os.environ.get("DSBENCH_DIR", "data/dsbench"))
OUTPUT_DIR = Path("output/dsbench_recognition")
CACHE_FILE = OUTPUT_DIR / "recognition_cache.json"
LOG_DIR = Path("logs")

# Create output and log directories
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)

# Setup logging with timestamp
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_file = LOG_DIR / f"dsbench_recognition_{timestamp}.log"

# File formatter for detailed logs
file_formatter = logging.Formatter(
    "%(asctime)s - [%(levelname)s] - %(name)s - %(message)s"
)

# Configure root logger
root_logger = logging.getLogger()
root_logger.setLevel(logging.DEBUG)
root_logger.handlers = []

# Rich console handler - INFO and above with rich formatting
rich_handler = RichHandler(
    console=console,
    level=logging.INFO,
    show_time=True,
    rich_tracebacks=True,
    tracebacks_show_locals=False,
)
root_logger.addHandler(rich_handler)

# File handler - DEBUG and above for all modules
file_handler = logging.FileHandler(log_file, encoding="utf-8")
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(file_formatter)
root_logger.addHandler(file_handler)

# Get logger for this module
logger = logging.getLogger(__name__)

# Suppress verbose SpreadsheetLLM logs in console (still logged to file)
logging.getLogger("spreadsheet_llm").setLevel(logging.WARNING)

logger.info(f"Logging to: {log_file}")


def load_cache() -> dict:
    """Load cached recognition results with nested structure: {file: {sheet: result}}."""
    if CACHE_FILE.exists():
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
                # Count total cached entries (files and sheets)
                total_sheets = sum(len(sheets) for sheets in cache.values())
                logger.info(
                    f"Loaded cache: {len(cache)} files, {total_sheets} sheet(s)"
                )
                return cache
        except Exception as e:
            logger.warning(f"Failed to load cache: {e}")
            return {}
    return {}


def save_cache(cache: dict):
    """Save recognition results to cache with nested structure: {file: {sheet: result}}."""
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2, ensure_ascii=False)
        total_sheets = sum(len(sheets) for sheets in cache.values())
        logger.debug(f"Saved cache: {len(cache)} files, {total_sheets} sheet(s)")
    except Exception as e:
        logger.error(f"Failed to save cache: {e}")


def find_excel_files(directory: Path) -> list[Path]:
    """Find all Excel files in directory recursively, excluding temp files."""
    excel_files = []

    # Find .xlsx files
    for file in directory.rglob("*.xlsx"):
        if not file.name.startswith("~$"):  # Exclude Excel temp files
            excel_files.append(file)

    # Find .xls files
    for file in directory.rglob("*.xls"):
        if not file.name.startswith("~$") and not file.name.endswith(".xlsx"):
            excel_files.append(file)

    return sorted(excel_files)


def create_annotated_file(
    file_path: Path,
    regions: list[dict],
    relative_path: Path,
    sheet_name: str | None = None,
) -> Optional[str]:
    """
    Create an annotated Excel file with colored regions.

    Args:
        file_path: Path to original Excel file
        regions: List of region dicts with 'title' and 'range' keys
        relative_path: Relative path from DSBENCH_DIR for preserving folder structure
        sheet_name: Name of the sheet to color (uses active sheet if None)

    Returns:
        Relative path to annotated file (from OUTPUT_DIR), or None if failed
    """
    try:
        logger.debug("Creating annotated file with colored regions...")

        # Load workbook
        import openpyxl

        wb_annotated = openpyxl.load_workbook(file_path)

        # Color identified regions with distinct background colors
        color_regions(wb_annotated, regions, sheet_name=sheet_name)

        # Save annotated workbook preserving directory structure
        annotated_dir = OUTPUT_DIR / "annotated"

        # Create output path: annotated/{relative_path}
        output_path = annotated_dir / relative_path

        # Ensure the subdirectory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb_annotated.save(output_path)
        logger.debug(f"Saved annotated file: {output_path}")

        return str(output_path.relative_to(OUTPUT_DIR))

    except Exception as e:
        logger.warning(f"Failed to create annotated file: {e}")
        logger.debug("Annotation error:", exc_info=True)
        return None


def color_regions(wb, regions: list[dict], sheet_name: str | None = None) -> None:
    """
    Color identified regions in the workbook with distinct background colors.

    Each region gets a unique color to visually distinguish different tables.
    All other cells will have their styles removed for better contrast.

    Args:
        wb: openpyxl Workbook object
        regions: List of region dicts with 'title' and 'range' keys
        sheet_name: Name of sheet to modify (uses active sheet if None)
    """
    from openpyxl.styles import Alignment, Border, Font

    # Get the worksheet
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Step 1: Clear all cell styles to make identified regions stand out
    logger.debug("Clearing all cell styles...")
    default_fill = PatternFill(fill_type=None)  # No fill
    default_font = Font()  # Default font
    default_alignment = Alignment()  # Default alignment
    default_border = Border()  # No borders

    for row in ws.iter_rows():
        for cell in row:
            cell.fill = default_fill
            cell.font = default_font
            cell.alignment = default_alignment
            cell.border = default_border

    # Define a palette of distinct colors (light/pastel colors for readability)
    color_palette = [
        "FFE699",  # Light yellow
        "B4C7E7",  # Light blue
        "C5E0B4",  # Light green
        "F8CBAD",  # Light orange
        "E2EFDA",  # Light mint
        "FCE4D6",  # Light peach
        "DDEBF7",  # Light sky blue
        "FFF2CC",  # Light cream
        "D9E1F2",  # Light lavender
        "E2F0D9",  # Light lime
        "FFDFD3",  # Light salmon
        "DAE3F3",  # Pale blue
    ]

    # Process each region with a different color
    for region_idx, region in enumerate(regions):
        # Cycle through colors if we have more regions than colors
        color = color_palette[region_idx % len(color_palette)]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        range_str = region["range"]

        # Handle multiple ranges separated by comma
        for single_range in range_str.split(","):
            single_range = single_range.strip()

            try:
                # Parse range
                if ":" in single_range:
                    start_col, start_row, end_col, end_row = parse_excel_range(
                        single_range
                    )
                else:
                    # Single cell
                    col_str = "".join(c for c in single_range if c.isalpha())
                    row_str = "".join(c for c in single_range if c.isdigit())
                    start_col = end_col = col_str
                    start_row = end_row = int(row_str)

                # Convert column letters to indices
                start_col_idx = col_to_index(start_col) + 1  # openpyxl is 1-indexed
                end_col_idx = col_to_index(end_col) + 1

                # Apply background color to all cells in the region
                for row in range(start_row, end_row + 1):
                    for col in range(start_col_idx, end_col_idx + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = fill

                logger.debug(
                    f"Colored region {region_idx + 1} ({single_range}) with {color}"
                )

            except Exception as e:
                logger.warning(f"Failed to color range '{single_range}': {e}")


def process_file(
    file_path: Path,
    wrapper: SpreadsheetLLMWrapper,
    model,
    format_aware: bool = True,
    sheet_name: str | int | None = None,
) -> Optional[dict]:
    """
    Process a single Excel file and identify table regions.

    Args:
        file_path: Path to Excel file
        wrapper: SpreadsheetLLM wrapper instance
        model: LangChain ChatModel for LLM recognition
        format_aware: Use format-aware compression
        sheet_name: Sheet to process (None for active sheet, int for index, str for name)

    Returns:
        Dictionary with recognition results or None if failed
    """
    try:
        logger.info(f"Processing: {file_path.name}")

        # Read spreadsheet
        logger.debug("Step 1: Reading spreadsheet...")
        wb = wrapper.read_spreadsheet(file_path)
        if wb is None:
            logger.warning(f"Failed to read: {file_path}")
            return None

        # Compress spreadsheet
        logger.debug("Step 2: Compressing spreadsheet...")
        result = wrapper.compress_spreadsheet(
            wb, format_aware=format_aware, sheet_name=sheet_name
        )
        if result is None:
            logger.warning(f"Failed to compress: {file_path}")
            return None

        # Log compression stats
        row_count = len(result.anchors.row_anchors)
        col_count = len(result.anchors.column_anchors)
        orig_rows, orig_cols = result.anchors.original_shape
        logger.debug(
            f"Compression: {orig_rows}x{orig_cols} → {row_count}x{col_count} anchors"
        )

        # LLM-based recognition
        logger.debug("Step 3: Running LLM recognition...")
        recognition_result = wrapper.recognize_original(
            compress_dict=result.compress_dict,
            sheet_compressor=result.sheet_compressor,
            model=model,
            user_prompt="Identify all table regions in this spreadsheet. Include headers, data areas, and summary sections.",
        )

        # Build result dictionary
        recognition_items = []
        for item in recognition_result.items:
            recognition_items.append(
                {
                    "title": item.title or "Untitled",
                    "range": item.range,
                }
            )

        result_dict = {
            "file": str(file_path.relative_to(DSBENCH_DIR)),
            "file_name": file_path.name,
            "sheet_name": result.sheet_name,
            "timestamp": datetime.now().isoformat(),
            "compression_stats": {
                "original_shape": list(result.anchors.original_shape),
                "compressed_shape": list(result.anchors.compressed_shape),
                "row_anchors": len(result.anchors.row_anchors),
                "col_anchors": len(result.anchors.column_anchors),
            },
            "recognition": {
                "reasoning": recognition_result.reasoning,
                "num_regions": len(recognition_items),
                "regions": recognition_items,
            },
        }

        logger.info(
            f"✓ Found {len(recognition_items)} regions in {file_path.name} (sheet: {result.sheet_name})"
        )

        return result_dict

    except Exception as e:
        logger.error(f"Error processing {file_path.name}: {e}")
        logger.debug("Full traceback:", exc_info=True)
        return None


def display_results_table(results: list[dict]):
    """
    Display summary table of results using Rich.

    Note: Uses console.print() for Rich Table formatting,
    which is not suitable for standard logging output.
    """
    table = Table(title=f"DSBench Recognition Results ({len(results)} files)")

    table.add_column("File", style="cyan", no_wrap=False)
    table.add_column("Sheet", style="white", no_wrap=False)
    table.add_column("Original Size", justify="right", style="yellow")
    table.add_column("Anchors", justify="right", style="blue")
    table.add_column("Regions", justify="right", style="green")
    table.add_column("Annotated", justify="center", style="magenta")

    for result in results:
        comp = result["compression_stats"]
        orig_shape = f"{comp['original_shape'][0]}×{comp['original_shape'][1]}"
        anchors = f"{comp['row_anchors']}×{comp['col_anchors']}"
        regions = str(result["recognition"]["num_regions"])
        has_annotation = "✓" if "annotated_file" in result else "✗"
        sheet_name = result.get("sheet_name", "N/A")

        table.add_row(
            result["file_name"],
            sheet_name,
            orig_shape,
            anchors,
            regions,
            has_annotation,
        )

    console.print("\n")
    console.print(table)


def main():
    """Main processing pipeline."""

    logger.info("=" * 70)
    logger.info("DSBENCH TABLE RECOGNITION PIPELINE")
    logger.info("=" * 70)
    logger.info(f"Source: {DSBENCH_DIR}")
    logger.info(f"Output: {OUTPUT_DIR}")

    # Load cache
    logger.info("Loading cache...")
    cache = load_cache()

    # Find Excel files
    logger.info("Finding Excel files...")
    excel_files = find_excel_files(DSBENCH_DIR)
    logger.info(f"Found {len(excel_files)} Excel files")

    if len(excel_files) == 0:
        logger.error("No Excel files found!")
        return

    # Initialize LLM model
    logger.info("Initializing LLM model...")
    try:
        from langchain_openai import ChatOpenAI

        model_name = os.environ.get("MODEL_NAME", "google/gemini-2.5-pro")
        model = ChatOpenAI(model=model_name)
        logger.info(f"Using model: {model_name}")
    except Exception as e:
        logger.error(f"Failed to initialize LLM: {e}")
        logger.error("Please set OPENAI_API_KEY environment variable")
        return

    # Initialize SpreadsheetLLM wrapper
    logger.info("Initializing SpreadsheetLLM...")
    wrapper = SpreadsheetLLMWrapper()

    # Process files with progress bar
    results = []
    processed_count = 0
    cached_count = 0
    failed_count = 0

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console,
    ) as progress:

        task = progress.add_task("[cyan]Processing files...", total=len(excel_files))

        for file_path in excel_files:
            rel_path = file_path.relative_to(DSBENCH_DIR)
            file_cache_key = str(rel_path)

            # First, process the file (or get from cache) to determine which sheet was used
            # For now, we use None (active sheet) as default
            sheet_to_process = None  # Could be made configurable in the future

            # Read workbook briefly to get the actual sheet name that will be processed
            wb_temp = wrapper.read_spreadsheet(file_path)
            if wb_temp is None:
                logger.warning(f"Failed to read: {file_path.name}")
                failed_count += 1
                progress.update(task, advance=1)
                continue

            # Determine the actual sheet name that will be used
            if sheet_to_process is None:
                if wb_temp.active is None:
                    logger.warning(f"No active sheet in: {file_path.name}")
                    failed_count += 1
                    progress.update(task, advance=1)
                    continue
                actual_sheet = wb_temp.active.title
            elif isinstance(sheet_to_process, int):
                actual_sheet = wb_temp.sheetnames[sheet_to_process]
            else:
                actual_sheet = sheet_to_process

            # Check cache with nested structure: cache[file][sheet]
            if file_cache_key in cache and actual_sheet in cache[file_cache_key]:
                result = cache[file_cache_key][actual_sheet]
                logger.info(f"[CACHED] {file_path.name} (sheet: {actual_sheet})")
                cached_count += 1
            else:
                # Process file with the determined sheet
                result = process_file(
                    file_path, wrapper, model, sheet_name=sheet_to_process
                )

                if result:
                    # Initialize file entry if it doesn't exist
                    if file_cache_key not in cache:
                        cache[file_cache_key] = {}

                    # Store result under the sheet name
                    cache[file_cache_key][actual_sheet] = result
                    save_cache(cache)
                    processed_count += 1
                else:
                    failed_count += 1
                    result = None

            # Generate annotated file (for both cached and newly processed files)
            if result:
                regions = result["recognition"]["regions"]
                sheet_name = result.get("sheet_name")  # Get sheet name if available
                annotated_path = create_annotated_file(
                    file_path, regions, rel_path, sheet_name=sheet_name
                )
                if annotated_path:
                    result["annotated_file"] = annotated_path

                results.append(result)

            progress.update(task, advance=1)

    # Display summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("PROCESSING COMPLETE")
    logger.info("=" * 70)
    logger.info(f"Processed: {processed_count} files")
    logger.info(f"Cached: {cached_count} files")
    logger.info(f"Failed: {failed_count} files")
    logger.info(f"Total successful: {len(results)} files")

    # Display results table
    if results:
        display_results_table(results)

        # Save detailed results
        output_file = OUTPUT_DIR / f"recognition_results_{timestamp}.json"
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "metadata": {
                        "timestamp": datetime.now().isoformat(),
                        "total_files": len(excel_files),
                        "successful": len(results),
                        "failed": failed_count,
                        "model": model_name,
                    },
                    "results": results,
                },
                f,
                indent=2,
                ensure_ascii=False,
            )

        logger.info("")
        logger.info(f"✓ Results saved to: {output_file}")

        # Calculate statistics
        total_regions = sum(r["recognition"]["num_regions"] for r in results)
        avg_regions = total_regions / len(results) if results else 0
        annotated_count = sum(1 for r in results if "annotated_file" in r)

        # Get unique sheet names
        sheet_names = set(r.get("sheet_name") for r in results if r.get("sheet_name"))

        logger.info("")
        logger.info("STATISTICS")
        logger.info("-" * 70)
        logger.info(f"Total regions identified: {total_regions}")
        logger.info(f"Average per file: {avg_regions:.1f}")
        logger.info(f"Annotated files: {annotated_count} / {len(results)}")
        if sheet_names:
            logger.info(f"Sheets processed: {', '.join(sorted(sheet_names))}")

        # Show annotated files location
        annotated_dir = OUTPUT_DIR / "annotated"
        if annotated_count > 0:
            logger.info("")
            logger.info(f"✓ Annotated files saved to: {annotated_dir}")
            logger.info("  (Colored backgrounds mark identified regions)")
            logger.info("  (Other cells have all formatting removed for contrast)")

    logger.info("")
    logger.info(f"Log file: {log_file}")
    logger.info("=" * 70)


if __name__ == "__main__":
    main()
