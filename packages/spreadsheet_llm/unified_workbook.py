"""
Unified interface for different Excel formats (xlsx, xlsb, xls).

Provides a consistent API regardless of the underlying library (openpyxl, pyxlsb, xlrd).
"""

from abc import ABC, abstractmethod
from typing import Any, Iterator, List, Optional

import openpyxl
import openpyxl.workbook
import pandas as pd
import pyxlsb


# Unified cell format classes that mimic openpyxl structure
class UnifiedBorderSide:
    """Unified border side (top/bottom/left/right)."""

    def __init__(self, style: Optional[str] = None):
        self.style = style


class UnifiedBorder:
    """Unified border object mimicking openpyxl.styles.Border."""

    def __init__(
        self,
        top: Optional[UnifiedBorderSide] = None,
        bottom: Optional[UnifiedBorderSide] = None,
        left: Optional[UnifiedBorderSide] = None,
        right: Optional[UnifiedBorderSide] = None,
    ):
        self.top = top or UnifiedBorderSide()
        self.bottom = bottom or UnifiedBorderSide()
        self.left = left or UnifiedBorderSide()
        self.right = right or UnifiedBorderSide()


class UnifiedFillColor:
    """Unified fill color object."""

    def __init__(self, index: str = "00000000"):
        self.index = index


class UnifiedFill:
    """Unified fill object mimicking openpyxl.styles.Fill."""

    def __init__(self, start_color: Optional[UnifiedFillColor] = None):
        self.start_color = start_color


class UnifiedFont:
    """Unified font object mimicking openpyxl.styles.Font."""

    def __init__(self, bold: Optional[bool] = None):
        self.bold = bold


class UnifiedCell(ABC):
    """Abstract base class for unified cell interface."""

    @property
    @abstractmethod
    def value(self) -> Any:
        """Get cell value."""
        pass

    @property
    @abstractmethod
    def border(self) -> UnifiedBorder:
        """Get cell border (mimics openpyxl)."""
        pass

    @property
    @abstractmethod
    def fill(self) -> UnifiedFill:
        """Get cell fill (mimics openpyxl)."""
        pass

    @property
    @abstractmethod
    def font(self) -> UnifiedFont:
        """Get cell font (mimics openpyxl)."""
        pass


class UnifiedWorksheet(ABC):
    """Abstract base class for unified worksheet interface."""

    @property
    @abstractmethod
    def title(self) -> str:
        """Get worksheet title/name."""
        pass

    @abstractmethod
    def cell(self, row: int, column: int) -> UnifiedCell:
        """
        Get cell at specified position.

        Args:
            row: 1-based row index
            column: 1-based column index

        Returns:
            UnifiedCell instance
        """
        pass

    @abstractmethod
    def iter_rows(
        self,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
    ) -> Iterator[tuple[UnifiedCell, ...]]:
        """
        Iterate over cells in the specified range.

        Args:
            min_row: Minimum row index (1-based, inclusive)
            max_row: Maximum row index (1-based, inclusive)
            min_col: Minimum column index (1-based, inclusive)
            max_col: Maximum column index (1-based, inclusive)

        Yields:
            Tuples of UnifiedCell instances for each row in the range
        """
        pass


class UnifiedWorkbook(ABC):
    """Abstract base class for unified workbook interface."""

    @property
    @abstractmethod
    def sheetnames(self) -> List[str]:
        """Get list of sheet names."""
        pass

    @property
    @abstractmethod
    def active(self) -> Optional[UnifiedWorksheet]:
        """Get active worksheet."""
        pass

    @property
    @abstractmethod
    def native_workbook(self) -> Any:
        """Get the underlying native workbook object (openpyxl, pyxlsb, etc.)."""
        pass

    @property
    @abstractmethod
    def excel_file(self) -> pd.ExcelFile:
        """Get the pandas ExcelFile object, if available."""
        pass

    @property
    @abstractmethod
    def file_path(self) -> Optional[str]:
        """Get the file path of the workbook, if available."""
        pass

    @abstractmethod
    def get_sheet_by_name(self, name: str) -> UnifiedWorksheet:
        """Get worksheet by name."""
        pass

    @abstractmethod
    def get_sheet_by_index(self, index: int) -> UnifiedWorksheet:
        """Get worksheet by index (0-based)."""
        pass

    def __getitem__(self, key: str) -> UnifiedWorksheet:
        """Get worksheet by name using dictionary syntax."""
        return self.get_sheet_by_name(key)

    @property
    def worksheets(self) -> List[UnifiedWorksheet]:
        """Get list of all worksheets."""
        return [self.get_sheet_by_index(i) for i in range(len(self.sheetnames))]


# ============================================================================
# OpenPyXL Implementation (for .xlsx)
# ============================================================================


class OpenpyxlCell(UnifiedCell):
    """Wrapper for openpyxl Cell. Wraps openpyxl objects with None checks."""

    def __init__(self, cell):
        self._cell = cell

    @property
    def value(self) -> Any:
        return self._cell.value

    @property
    def border(self):
        """Return openpyxl border with guaranteed non-None sides."""
        native_border = self._cell.border
        if native_border is None:
            return UnifiedBorder()

        # Wrap openpyxl border, ensuring all sides are safe to access
        # openpyxl may have None for individual sides (top/bottom/left/right)
        return _wrap_openpyxl_border(native_border)

    @property
    def fill(self):
        """Return openpyxl fill, or empty fill if None."""
        return self._cell.fill if self._cell.fill is not None else UnifiedFill()

    @property
    def font(self):
        """Return openpyxl font, or empty font if None."""
        return self._cell.font if self._cell.font is not None else UnifiedFont()


def _wrap_openpyxl_border(native_border):
    """
    Wrap openpyxl Border with guaranteed non-None sides.

    openpyxl may have None for individual border sides (top/bottom/left/right).
    This ensures all sides are UnifiedBorderSide objects.
    """
    return UnifiedBorder(
        top=UnifiedBorderSide(native_border.top.style if native_border.top else None),
        bottom=UnifiedBorderSide(
            native_border.bottom.style if native_border.bottom else None
        ),
        left=UnifiedBorderSide(
            native_border.left.style if native_border.left else None
        ),
        right=UnifiedBorderSide(
            native_border.right.style if native_border.right else None
        ),
    )


class OpenpyxlWorksheet(UnifiedWorksheet):
    """Wrapper for openpyxl Worksheet."""

    def __init__(self, worksheet):
        self._ws = worksheet

    @property
    def title(self) -> str:
        return self._ws.title

    def cell(self, row: int, column: int) -> UnifiedCell:
        return OpenpyxlCell(self._ws.cell(row=row, column=column))

    def iter_rows(
        self,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
    ) -> Iterator[tuple[OpenpyxlCell, ...]]:
        for row_cells in self._ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            yield tuple(OpenpyxlCell(cell) for cell in row_cells)


class OpenpyxlWorkbook(UnifiedWorkbook):
    """Wrapper for openpyxl Workbook."""

    def __init__(
        self,
        workbook: openpyxl.Workbook,
        excel_file: Optional[pd.ExcelFile] = None,
        file_path: Optional[str] = None,
    ):
        self._wb = workbook
        self._excel_file = excel_file
        self._file_path = file_path

    @property
    def sheetnames(self) -> List[str]:
        return self._wb.sheetnames

    @property
    def active(self) -> Optional[UnifiedWorksheet]:
        ws = self._wb.active
        return OpenpyxlWorksheet(ws) if ws else None

    @property
    def native_workbook(self) -> Any:
        return self._wb

    @property
    def excel_file(self) -> Optional[pd.ExcelFile]:
        return self._excel_file

    @property
    def file_path(self) -> Optional[str]:
        return self._file_path

    def get_sheet_by_name(self, name: str) -> UnifiedWorksheet:
        return OpenpyxlWorksheet(self._wb[name])

    def get_sheet_by_index(self, index: int) -> UnifiedWorksheet:
        return OpenpyxlWorksheet(self._wb.worksheets[index])


# ============================================================================
# PyXLSB Implementation (for .xlsb)
# ============================================================================


class PyxlsbCell(UnifiedCell):
    """
    Wrapper for pyxlsb Cell.

    Note: pyxlsb doesn't provide cell objects with format info.
    We only have access to the cell value. All format properties return None/empty.
    """

    def __init__(self, cell_value: Any):
        self._value = cell_value
        # Create empty format objects
        self._border = UnifiedBorder()
        self._fill = UnifiedFill()
        self._font = UnifiedFont()

    @property
    def value(self) -> Any:
        return self._value

    @property
    def border(self) -> UnifiedBorder:
        """Return empty border (xlsb doesn't support format info)."""
        return self._border

    @property
    def fill(self) -> UnifiedFill:
        """Return empty fill (xlsb doesn't support format info)."""
        return self._fill

    @property
    def font(self) -> UnifiedFont:
        """Return empty font (xlsb doesn't support format info)."""
        return self._font


class PyxlsbWorksheet(UnifiedWorksheet):
    """Wrapper for pyxlsb Worksheet."""

    def __init__(self, workbook: pyxlsb.Workbook, sheet_name: str):
        self._wb = workbook
        self._sheet_name = sheet_name
        # Cache the sheet data for performance
        self._sheet_data = None

    def _load_sheet_data(self):
        """Lazy load sheet data into memory."""
        if self._sheet_data is None:
            # pyxlsb returns generator, convert to list for random access
            with self._wb.get_sheet(self._sheet_name) as sheet:
                self._sheet_data = list(sheet.rows())

    @property
    def title(self) -> str:
        return self._sheet_name

    def cell(self, row: int, column: int) -> UnifiedCell:
        """
        Get cell value at position (1-based indexing).

        Note: pyxlsb uses 0-based indexing internally.
        """
        self._load_sheet_data()

        # Convert to 0-based
        row_idx = row - 1
        col_idx = column - 1

        if self._sheet_data is None:
            raise ValueError("Sheet data not loaded")

        # Check bounds
        if row_idx < 0 or row_idx >= len(self._sheet_data):
            return PyxlsbCell(None)

        row_data = self._sheet_data[row_idx]
        if col_idx < 0 or col_idx >= len(row_data):
            return PyxlsbCell(None)

        cell_data = row_data[col_idx]
        # pyxlsb returns Cell objects with .v attribute for value
        value = cell_data.v if hasattr(cell_data, "v") else cell_data
        return PyxlsbCell(value)

    def iter_rows(
        self,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
    ) -> Iterator[tuple[PyxlsbCell, ...]]:
        self._load_sheet_data()

        if self._sheet_data is None:
            return

        # Determine the actual range
        total_rows = len(self._sheet_data)
        start_row = (min_row or 1) - 1  # Convert to 0-based
        end_row = (max_row or total_rows) - 1  # Convert to 0-based

        # Clamp to valid range
        start_row = max(0, min(start_row, total_rows - 1))
        end_row = max(0, min(end_row, total_rows - 1))

        for row_idx in range(start_row, end_row + 1):
            row_data = self._sheet_data[row_idx]
            total_cols = len(row_data)

            start_col = (min_col or 1) - 1  # Convert to 0-based
            end_col = (max_col or total_cols) - 1  # Convert to 0-based

            # Clamp to valid range
            start_col = max(0, min(start_col, total_cols - 1)) if total_cols > 0 else 0
            end_col = max(0, min(end_col, total_cols - 1)) if total_cols > 0 else 0

            # Generate cells for this row
            cells = []
            for col_idx in range(start_col, end_col + 1):
                if col_idx < len(row_data):
                    cell_data = row_data[col_idx]
                    value = cell_data.v if hasattr(cell_data, "v") else cell_data
                    cells.append(PyxlsbCell(value))
                else:
                    cells.append(PyxlsbCell(None))

            yield tuple(cells)


class PyxlsbWorkbook(UnifiedWorkbook):
    """Wrapper for pyxlsb Workbook."""

    def __init__(
        self,
        workbook: pyxlsb.Workbook,
        excel_file: Optional[pd.ExcelFile] = None,
        file_path: Optional[str] = None,
    ):
        self._wb = workbook
        self._excel_file = excel_file
        self._file_path = file_path

    @property
    def sheetnames(self) -> List[str]:
        sheets = self._wb.sheets
        return sheets if sheets is not None else []

    @property
    def active(self) -> Optional[UnifiedWorksheet]:
        # pyxlsb doesn't have concept of "active" sheet
        # Return first sheet by convention
        sheet_list = self.sheetnames
        if sheet_list:
            return self.get_sheet_by_index(0)
        return None

    @property
    def native_workbook(self) -> Any:
        return self._wb

    @property
    def excel_file(self) -> Optional[pd.ExcelFile]:
        return self._excel_file

    @property
    def file_path(self) -> Optional[str]:
        return self._file_path

    def get_sheet_by_name(self, name: str) -> UnifiedWorksheet:
        if name not in self.sheetnames:
            raise KeyError(f"Sheet '{name}' not found")
        return PyxlsbWorksheet(self._wb, name)

    def get_sheet_by_index(self, index: int) -> UnifiedWorksheet:
        sheet_list = self.sheetnames
        sheet_name = sheet_list[index]
        return PyxlsbWorksheet(self._wb, sheet_name)


# ============================================================================
# Factory Function
# ============================================================================


def create_unified_workbook(file_path: str) -> UnifiedWorkbook:
    """
    Create a UnifiedWorkbook from an Excel file.

    Automatically detects file format and returns appropriate wrapper.

    Args:
        file_path: Path to Excel file (.xlsx, .xlsb, .xls, etc.)

    Returns:
        UnifiedWorkbook instance

    Example:
        >>> wb = create_unified_workbook("data.xlsx")
        >>> ws = wb.active
        >>> cell = ws.cell(1, 1)
        >>> print(cell.value)
    """
    excel_file = pd.ExcelFile(file_path)
    book = excel_file.book

    # Detect book type and wrap accordingly
    if isinstance(book, openpyxl.workbook.workbook.Workbook):
        return OpenpyxlWorkbook(book, excel_file=excel_file, file_path=file_path)
    elif isinstance(book, pyxlsb.Workbook):
        return PyxlsbWorkbook(book, excel_file=excel_file, file_path=file_path)
    else:
        # For xlrd or other unsupported types
        book_type = type(book).__module__ + "." + type(book).__name__
        raise ValueError(f"Unsupported workbook type: {book_type}")


def create_unified_workbook_from_excel_file(
    excel_file: pd.ExcelFile,
) -> UnifiedWorkbook:
    """
    Create a UnifiedWorkbook from a pandas ExcelFile.

    Args:
        excel_file: pandas ExcelFile instance

    Returns:
        UnifiedWorkbook instance
    """
    book = excel_file.book

    # Detect book type and wrap accordingly
    if isinstance(book, openpyxl.workbook.workbook.Workbook):
        return OpenpyxlWorkbook(book, excel_file=excel_file)
    elif isinstance(book, pyxlsb.Workbook):
        return PyxlsbWorkbook(book, excel_file=excel_file)
    else:
        # For xlrd or other unsupported types
        book_type = type(book).__module__ + "." + type(book).__name__
        raise ValueError(f"Unsupported workbook type: {book_type}")
